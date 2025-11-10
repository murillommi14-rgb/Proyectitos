# vistas del módulo, usamos genéricas para no escribir de más
from django.urls import reverse_lazy
from django.views.generic import ListView, CreateView, UpdateView, DeleteView, FormView
from django.http import HttpResponse, JsonResponse
from django.db.models import Q
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
import openpyxl
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
import pandas as pd
from datetime import datetime
import re
import pdfplumber
from .models import Calificacion
from .forms import CalificacionForm, BulkUploadForm

@method_decorator(login_required, name='dispatch')
class CalificacionListView(ListView):
    # lista con paginación
    model = Calificacion
    paginate_by = 10
    template_name = "calificaciones/lista.html"

    def get_queryset(self):
        queryset = super().get_queryset()
        tipo_mercado = self.request.GET.get('tipo_mercado')
        origen = self.request.GET.get('origen')
        ejercicio = self.request.GET.get('ejercicio')
        mercado = self.request.GET.get('mercado')
        if tipo_mercado:
            queryset = queryset.filter(tipo_mercado=tipo_mercado)
        if origen:
            queryset = queryset.filter(origen=origen)
        if ejercicio:
            queryset = queryset.filter(ejercicio=ejercicio)
        if mercado:
            queryset = queryset.filter(mercado__icontains=mercado)
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filter_tipo_mercado'] = self.request.GET.get('tipo_mercado', '')
        context['filter_origen'] = self.request.GET.get('origen', '')
        context['filter_ejercicio'] = self.request.GET.get('ejercicio', '')
        context['filter_mercado'] = self.request.GET.get('mercado', '')
        context['form'] = CalificacionForm()
        return context

    def get(self, request, *args, **kwargs):
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest' and 'edit' in request.GET:
            calificacion_id = request.GET.get('edit')
            try:
                calificacion = Calificacion.objects.get(pk=calificacion_id)
                data = {
                    'success': True,
                    'calificacion': {
                        'estado': calificacion.estado,
                        'secuencia': calificacion.secuencia,
                        'tipo_mercado': calificacion.tipo_mercado,
                        'origen': calificacion.origen,
                        'mercado': calificacion.mercado,
                        'instrumento': calificacion.instrumento,
                        'numero_dividendo': calificacion.numero_dividendo,
                        'tipo_sociedad': calificacion.tipo_sociedad,
                        'valor_historico': float(calificacion.valor_historico),
                        'factor_actualizacion': float(calificacion.factor_actualizacion),
                        'fecha': calificacion.fecha.strftime('%Y-%m-%d') if calificacion.fecha else '',
                        'ejercicio': calificacion.ejercicio.strftime('%Y-%m-%d') if calificacion.ejercicio else '',
                        'isfut_casilla': calificacion.isfut_casilla,
                        'descripcion': calificacion.descripcion,
                        'factor1': float(calificacion.factor1 or 0),
                        'factor2': float(calificacion.factor2 or 0),
                        'factor3': float(calificacion.factor3 or 0),
                        'factor4': float(calificacion.factor4 or 0),
                        'factor5': float(calificacion.factor5 or 0),
                        'factor6': float(calificacion.factor6 or 0),
                        'factor7': float(calificacion.factor7 or 0),
                        'factor8': float(calificacion.factor8 or 0),
                        'factor9': float(calificacion.factor9 or 0),
                        'factor10': float(calificacion.factor10 or 0),
                        'factor11': float(calificacion.factor11 or 0),
                        'factor12': float(calificacion.factor12 or 0),
                        'factor13': float(calificacion.factor13 or 0),
                        'factor14': float(calificacion.factor14 or 0),
                        'factor15': float(calificacion.factor15 or 0),
                        'factor16': float(calificacion.factor16 or 0),
                        'factor17': float(calificacion.factor17 or 0),
                        'factor18': float(calificacion.factor18 or 0),
                        'factor19': float(calificacion.factor19 or 0),
                        'factor20': float(calificacion.factor20 or 0),
                        'factor21': float(calificacion.factor21 or 0),
                        'factor22': float(calificacion.factor22 or 0),
                        'factor23': float(calificacion.factor23 or 0),
                        'factor24': float(calificacion.factor24 or 0),
                        'factor25': float(calificacion.factor25 or 0),
                        'factor26': float(calificacion.factor26 or 0),
                    }
                }
                return JsonResponse(data)
            except Calificacion.DoesNotExist:
                return JsonResponse({'success': False, 'error': 'Calificación no encontrada'})
        return super().get(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            # Manejar crear
            form = CalificacionForm(request.POST)
            if form.is_valid():
                try:
                    calificacion = form.save()
                    # Devuelve los nuevos datos de calificación para la actualización dinámica de la tabla.
                    data = {
                        'success': True,
                        'message': 'Calificación creada exitosamente.',
                        'calificacion': {
                            'id': calificacion.pk,
                            'tipo_mercado_display': calificacion.get_tipo_mercado_display(),
                            'origen_display': calificacion.get_origen_display(),
                            'mercado': calificacion.mercado,
                            'instrumento': calificacion.instrumento,
                            'secuencia': calificacion.secuencia,
                            'numero_dividendo': calificacion.numero_dividendo,
                            'tipo_sociedad': calificacion.tipo_sociedad,
                            'fecha': calificacion.fecha.strftime('%d/m/Y') if calificacion.fecha else '',
                            'valor_historico': f"${calificacion.valor_historico:,}",
                            'ejercicio': calificacion.ejercicio.strftime('%Y') if calificacion.ejercicio else '',
                            'descripcion': calificacion.descripcion[:50] + '...' if len(calificacion.descripcion) > 50 else calificacion.descripcion,
                            'isfut_casilla': 'Sí' if calificacion.isfut_casilla else 'No',
                            'factor_actualizacion': f"{calificacion.factor_actualizacion:.4f}",
                            'estado': calificacion.estado,
                            'estado_badge': f'<span class="badge bg-{"success" if calificacion.estado == "VIGENTE" else "danger" if calificacion.estado == "ANULADA" else "warning"}">{calificacion.estado.title()}</span>',
                            'actualizado_en': calificacion.actualizado_en.strftime('%d/m/Y H:i'),
                        }
                    }
                    return JsonResponse(data)
                except Exception as e:
                    return JsonResponse({'success': False, 'error': str(e)})
            else:
                errors = []
                for field, error_list in form.errors.items():
                    for error in error_list:
                        errors.append(f"{field}: {error}")
                return JsonResponse({'success': False, 'error': ' '.join(errors)})
        # Para solicitudes que no sean AJAX, redirija a la vista de lista.
        from django.http import HttpResponseRedirect
        return HttpResponseRedirect(reverse_lazy('calif_list'))

@method_decorator(login_required, name='dispatch')
class CalificacionCreateView(CreateView):
    # formulario de creación (valida reglas en model.clean())
    model = Calificacion
    form_class = CalificacionForm
    success_url = reverse_lazy("calif_list")
    template_name = "calificaciones/crear.html"

    def form_valid(self, form):
        try:
            self.object = form.save()
            if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': True, 'message': 'Calificación creada exitosamente.'})
            messages.success(self.request, 'Calificación creada exitosamente.')
            return super().form_valid(form)
        except Exception as e:
            if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'error': str(e)})
            messages.error(self.request, f'Error al crear la calificación: {str(e)}')
            return self.form_invalid(form)

    def form_invalid(self, form):
        if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            errors = []
            for field, error_list in form.errors.items():
                for error in error_list:
                    errors.append(f"{field}: {error}")
            return JsonResponse({'success': False, 'error': ' '.join(errors)})
        return super().form_invalid(form)


@method_decorator(login_required, name='dispatch')
class CalificacionUpdateView(UpdateView):
    # formulario de edición
    model = Calificacion
    form_class = CalificacionForm
    success_url = reverse_lazy("calif_list")
    template_name = "calificaciones/crear.html"

    def form_valid(self, form):
        try:
            self.object = form.save()
            if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                # Devuelve los datos de calificación actualizados para la actualización dinámica de la tabla.
                data = {
                    'success': True,
                    'message': 'Calificación actualizada exitosamente.',
                    'calificacion': {
                        'id': self.object.pk,
                        'tipo_mercado_display': self.object.get_tipo_mercado_display(),
                        'origen_display': self.object.get_origen_display(),
                        'mercado': self.object.mercado,
                        'instrumento': self.object.instrumento,
                        'secuencia': self.object.secuencia,
                        'numero_dividendo': self.object.numero_dividendo,
                        'tipo_sociedad': self.object.tipo_sociedad,
                        'fecha': self.object.fecha.strftime('%d/%m/%Y') if self.object.fecha else '',
                        'valor_historico': f"${self.object.valor_historico:,}",
                        'ejercicio': self.object.ejercicio.strftime('%Y') if self.object.ejercicio else '',
                        'descripcion': self.object.descripcion[:50] + '...' if len(self.object.descripcion) > 50 else self.object.descripcion,
                        'isfut_casilla': 'Sí' if self.object.isfut_casilla else 'No',
                        'factor_actualizacion': f"{self.object.factor_actualizacion:.4f}",
                        'estado': self.object.estado,
                        'estado_badge': f'<span class="badge bg-{"success" if self.object.estado == "VIGENTE" else "danger" if self.object.estado == "ANULADA" else "warning"}">{self.object.estado.title()}</span>',
                        'actualizado_en': self.object.actualizado_en.strftime('%d/m/Y H:i'),
                    }
                }
                return JsonResponse(data)
            messages.success(self.request, 'Calificación actualizada exitosamente.')
            return super().form_valid(form)
        except Exception as e:
            if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'error': str(e)})
            messages.error(self.request, f'Error al actualizar la calificación: {str(e)}')
            return self.form_invalid(form)

    def form_invalid(self, form):
        if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            errors = []
            for field, error_list in form.errors.items():
                for error in error_list:
                    errors.append(f"{field}: {error}")
            return JsonResponse({'success': False, 'error': ' '.join(errors)})
        return super().form_invalid(form)


@method_decorator(login_required, name='dispatch')
class CalificacionDeleteView(DeleteView):
    # vista de eliminación
    model = Calificacion
    success_url = reverse_lazy("calif_list")
    template_name = "calificaciones/confirm_delete.html"


@login_required
def download_template_csv(request):
    import csv
    from django.http import HttpResponse

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="plantilla_calificaciones.csv"'

    writer = csv.writer(response)
    headers = ["tipo_mercado", "origen", "mercado", "instrumento", "secuencia", "numero_dividendo", "tipo_sociedad", "fecha", "valor_historico", "ejercicio", "descripcion", "isfut_casilla", "factor_actualizacion", "factor1", "factor2", "factor3", "factor4", "factor5", "factor6", "factor7", "factor8", "factor9", "factor10", "factor11", "factor12", "factor13", "factor14", "factor15", "factor16", "factor17", "factor18", "factor19", "factor20", "factor21", "factor22", "factor23", "factor24", "factor25", "factor26", "estado"]
    writer.writerow(headers)
    # Agregar una fila de muestra
    sample = ["ACCIONES", "NACIONAL", "Nacional", "Ejemplo Instrumento", 1, 1, "A", "2023-01-01", 1000.00, 2023, "Ejemplo Descripción", False, 1.0, 0.0385, 0.0385, 0.0385, 0.0385, 0.0385, 0.0385, 0.0385, 0.0385, 0.0385, 0.0385, 0.0385, 0.0385, 0.0385, 0.0385, 0.0385, 0.0385, 0.0385, 0.0385, 0.0385, 0.0385, 0.0385, 0.0385, 0.0385, 0.0385, 0.0385, 0.0385, 0.0385, "BORRADOR"]
    writer.writerow(sample)

    return response

@login_required
def download_template_xlsx(request):
    import openpyxl
    from django.http import HttpResponse

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Plantilla Calificaciones"

    headers = ["tipo_mercado", "origen", "mercado", "instrumento", "secuencia", "numero_dividendo", "tipo_sociedad", "fecha", "valor_historico", "ejercicio", "descripcion", "isfut_casilla", "factor_actualizacion", "factor1", "factor2", "factor3", "factor4", "factor5", "factor6", "factor7", "factor8", "factor9", "factor10", "factor11", "factor12", "factor13", "factor14", "factor15", "factor16", "factor17", "factor18", "factor19", "factor20", "factor21", "factor22", "factor23", "factor24", "factor25", "factor26", "estado"]
    ws.append(headers)

    # Agregar una fila de muestra
    sample = ["ACCIONES", "NACIONAL", "Nacional", "Ejemplo Instrumento", 1, 1, "A", "2023-01-01", 1000.00, 2023, "Ejemplo Descripción", False, 1.0, 0.0385, 0.0385, 0.0385, 0.0385, 0.0385, 0.0385, 0.0385, 0.0385, 0.0385, 0.0385, 0.0385, 0.0385, 0.0385, 0.0385, 0.0385, 0.0385, 0.0385, 0.0385, 0.0385, 0.0385, 0.0385, 0.0385, 0.0385, 0.0385, 0.0385, 0.0385, 0.0385, "BORRADOR"]
    ws.append(sample)

    # Formación
    from openpyxl.styles import Font, Alignment
    header_font = Font(bold=True)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="plantilla_calificaciones.xlsx"'
    wb.save(response)

    return response

@login_required
def download_template_pdf(request):
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
    from reportlab.lib import colors
    from django.http import HttpResponse

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="plantilla_calificaciones.pdf"'

    doc = SimpleDocTemplate(response, pagesize=letter)
    elements = []

    # Datos simples
    data = [
        ["tipo_mercado", "origen", "mercado", "instrumento", "secuencia", "numero_dividendo", "tipo_sociedad", "fecha", "valor_historico", "ejercicio", "descripcion", "isfut_casilla", "factor_actualizacion", "factor1", "factor2", "factor3", "factor4", "factor5", "factor6", "factor7", "factor8", "factor9", "factor10", "factor11", "factor12", "factor13", "factor14", "factor15", "factor16", "factor17", "factor18", "factor19", "factor20", "factor21", "factor22", "factor23", "factor24", "factor25", "factor26", "estado"],
        ["ACCIONES", "NACIONAL", "Nacional", "Ejemplo Instrumento", 1, 1, "A", "2023-01-01", 1000.00, 2023, "Ejemplo Descripción", False, 1.0, 0.0385, 0.0385, 0.0385, 0.0385, 0.0385, 0.0385, 0.0385, 0.0385, 0.0385, 0.0385, 0.0385, 0.0385, 0.0385, 0.0385, 0.0385, 0.0385, 0.0385, 0.0385, 0.0385, 0.0385, 0.0385, 0.0385, 0.0385, 0.0385, 0.0385, 0.0385, 0.0385, "BORRADOR"]
    ]

    # Crear Tabla
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))

    elements.append(table)
    doc.build(elements)

    return response

@login_required
def export_excel(request):
    queryset = Calificacion.objects.all()
    tipo_mercado = request.GET.get('tipo_mercado')
    origen = request.GET.get('origen')
    ejercicio = request.GET.get('ejercicio')
    mercado = request.GET.get('mercado')
    if tipo_mercado:
        queryset = queryset.filter(tipo_mercado=tipo_mercado)
    if origen:
        queryset = queryset.filter(origen=origen)
    if ejercicio:
        queryset = queryset.filter(ejercicio=ejercicio)
    if mercado:
        queryset = queryset.filter(mercado__icontains=mercado)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Calificaciones NUAM"
    headers = ["Tipo Mercado", "Origen", "Mercado", "Instrumento", "Secuencia", "Número Dividendo", "Tipo Sociedad", "Fecha", "Valor Histórico", "Ejercicio", "Descripción", "ISFUT", "Factor Actualización", "Estado"]
    ws.append(headers)
    for c in queryset:
        ws.append([
            c.get_tipo_mercado_display(), c.get_origen_display(), c.mercado, c.instrumento, c.secuencia, c.numero_dividendo, c.tipo_sociedad, c.fecha.strftime("%d-%m-%Y") if c.fecha else "", float(c.valor_historico), c.ejercicio, c.descripcion, c.isfut_casilla, float(c.factor_actualizacion), c.estado
        ])

    # Formación
    from openpyxl.styles import Font, Alignment
    header_font = Font(bold=True)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=calificaciones_nuam.xlsx'
    wb.save(response)
    return response


@login_required
def export_pdf(request):
    queryset = Calificacion.objects.all()
    tipo_mercado = request.GET.get('tipo_mercado')
    origen = request.GET.get('origen')
    ejercicio = request.GET.get('ejercicio')
    mercado = request.GET.get('mercado')
    if tipo_mercado:
        queryset = queryset.filter(tipo_mercado=tipo_mercado)
    if origen:
        queryset = queryset.filter(origen=origen)
    if ejercicio:
        queryset = queryset.filter(ejercicio=ejercicio)
    if mercado:
        queryset = queryset.filter(mercado__icontains=mercado)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=calificaciones_nuam.pdf'
    p = canvas.Canvas(response, pagesize=letter)
    width, height = letter

    # Título
    p.setFont("Helvetica-Bold", 16)
    p.drawString(200, height - 50, "Calificaciones Tributarias - NUAM")
    p.setFont("Helvetica", 12)
    p.drawString(50, height - 80, f"Total registros: {queryset.count()}")

    y = height - 100
    p.setFont("Helvetica-Bold", 10)
    p.drawString(50, y, "Ejercicio")
    p.drawString(120, y, "Mercado")
    p.drawString(200, y, "Instrumento")
    p.drawString(300, y, "Fecha")
    p.drawString(380, y, "Estado")
    y -= 15

    p.setFont("Helvetica", 9)
    for c in queryset:
        if y < 50:
            p.showPage()
            y = height - 50
            p.setFont("Helvetica-Bold", 10)
            p.drawString(50, y, "Ejercicio")
            p.drawString(120, y, "Mercado")
            p.drawString(200, y, "Instrumento")
            p.drawString(300, y, "Fecha")
            p.drawString(380, y, "Estado")
            y -= 15
            p.setFont("Helvetica", 9)
        p.drawString(50, y, str(c.ejercicio))
        p.drawString(120, y, str(c.mercado)[:20])
        p.drawString(200, y, str(c.instrumento)[:30])
        p.drawString(300, y, c.fecha.strftime("%d-%m-%Y") if c.fecha else "")
        p.drawString(380, y, c.estado)
        y -= 12

    p.save()
    return response


@login_required
def preview_bulk_upload(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido'})

    files = request.FILES.getlist('file')
    if not files:
        return JsonResponse({'error': 'No se proporcionaron archivos'})

    previews = []
    for file in files:
        try:
            if file.name.endswith('.csv'):
                df = pd.read_csv(file, encoding='utf-8')
            elif file.name.endswith(('.xlsx', '.xls')):
                df = pd.read_excel(file)
            elif file.name.endswith('.pdf'):
                with pdfplumber.open(file) as pdf:
                    page = pdf.pages[0]
                    table = page.extract_table()
                    if table:
                        df = pd.DataFrame(table[1:], columns=table[0])
                    else:
                        previews.append({'filename': file.name, 'error': 'No se encontró tabla en el PDF'})
                        continue
            else:
                previews.append({'filename': file.name, 'error': 'Formato de archivo no soportado'})
                continue

            # Validar columnas requeridas
            required_columns = ['tipo_mercado', 'origen', 'mercado', 'instrumento', 'secuencia', 'numero_dividendo', 'tipo_sociedad', 'fecha', 'valor_historico', 'ejercicio', 'descripcion', 'isfut_casilla', 'factor_actualizacion', 'factor1', 'factor2', 'factor3', 'factor4', 'factor5', 'factor6', 'factor7', 'factor8', 'factor9', 'factor10', 'factor11', 'factor12', 'factor13', 'factor14', 'factor15', 'factor16', 'factor17', 'factor18', 'factor19', 'factor20', 'factor21', 'factor22', 'factor23', 'factor24', 'factor25', 'factor26', 'estado']
            if not all(col in df.columns for col in required_columns):
                previews.append({'filename': file.name, 'error': 'El archivo no tiene las columnas requeridas'})
                continue

            # Tomar primeras 5 filas para previsualización
            preview_df = df.head(5)
            headers = list(preview_df.columns)
            rows = preview_df.values.tolist()

            previews.append({
                'filename': file.name,
                'headers': headers,
                'rows': rows
            })

        except Exception as e:
            previews.append({'filename': file.name, 'error': f'Error al procesar el archivo: {str(e)}'})

    return JsonResponse({'previews': previews})

@login_required
def bulk_upload_ajax(request):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método no permitido'})

    files = request.FILES.getlist('file')
    total_calificaciones = 0
    errors = []

    for file in files:
        try:
            if file.name.endswith('.csv'):
                df = pd.read_csv(file, encoding='utf-8')
            elif file.name.endswith(('.xlsx', '.xls')):
                df = pd.read_excel(file)
            elif file.name.endswith('.pdf'):
                with pdfplumber.open(file) as pdf:
                    page = pdf.pages[0]
                    table = page.extract_table()
                    if table:
                        df = pd.DataFrame(table[1:], columns=table[0])
                    else:
                        errors.append(f"No se encontró tabla en el PDF {file.name}.")
                        continue
            else:
                errors.append(f"Formato de archivo no soportado para {file.name}. Usa CSV, XLSX o PDF.")
                continue

            # Validar columnas requeridas
            required_columns = ['tipo_mercado', 'origen', 'mercado', 'instrumento', 'secuencia', 'numero_dividendo', 'tipo_sociedad', 'fecha', 'valor_historico', 'ejercicio', 'descripcion', 'isfut_casilla', 'factor_actualizacion', 'factor1', 'factor2', 'factor3', 'factor4', 'factor5', 'factor6', 'factor7', 'factor8', 'factor9', 'factor10', 'factor11', 'factor12', 'factor13', 'factor14', 'factor15', 'factor16', 'factor17', 'factor18', 'factor19', 'factor20', 'factor21', 'factor22', 'factor23', 'factor24', 'factor25', 'factor26', 'estado']
            if not all(col in df.columns for col in required_columns):
                errors.append(f"El archivo {file.name} no tiene las columnas requeridas.")
                continue

            calificaciones = []
            for _, row in df.iterrows():
                try:
                    tipo_mercado_val = str(row['tipo_mercado']).strip() if row['tipo_mercado'] else 'ACCIONES'
                    origen_val = str(row['origen']).strip() if row['origen'] else 'CORREDORA'
                    mercado_val = str(row['mercado']).strip()
                    instrumento_val = str(row['instrumento']).strip()
                    secuencia_val = int(float(row['secuencia'])) if row['secuencia'] and str(row['secuencia']).strip() else 0
                    numero_dividendo_val = int(float(row['numero_dividendo'])) if row['numero_dividendo'] and str(row['numero_dividendo']).strip() else 0
                    tipo_sociedad_val = str(row['tipo_sociedad']).strip() if row['tipo_sociedad'] else 'A'
                    if tipo_sociedad_val not in ['A', 'C']:
                        tipo_sociedad_val = 'A'
                    fecha_val = pd.to_datetime(row['fecha'], errors='coerce') if row['fecha'] else datetime(2023, 1, 1)
                    if pd.isna(fecha_val):
                        fecha_val = datetime(2023, 1, 1)
                    fecha_val = fecha_val.date()
                    valor_historico_val = float(row['valor_historico']) if row['valor_historico'] else 0.0
                    ejercicio_val = pd.to_datetime(row['ejercicio'], errors='coerce') if row['ejercicio'] else datetime(2023, 1, 1)
                    if pd.isna(ejercicio_val):
                        ejercicio_val = datetime(2023, 1, 1)
                    ejercicio_val = ejercicio_val.date()
                    descripcion_val = str(row['descripcion']).strip() if row['descripcion'] else ''
                    # Convertir isfut_casilla a booleano
                    isfut_casilla_val = str(row['isfut_casilla']).strip().lower() if row['isfut_casilla'] else ''
                    if isfut_casilla_val in ['true', 'verdadero', '1', 'yes', 'si']:
                        isfut_casilla_val = True
                    elif isfut_casilla_val in ['false', 'falso', '0', 'no']:
                        isfut_casilla_val = False
                    else:
                        isfut_casilla_val = False  # Valor predeterminado: Falso si no se reconoce
                    factor_actualizacion_val = float(row['factor_actualizacion']) if row['factor_actualizacion'] else 1.0
                    estado_val = str(row.get('estado', 'BORRADOR')).strip() or 'BORRADOR'
                    if estado_val not in ['BORRADOR', 'EN PROCESO', 'VIGENTE', 'ANULADA', 'INVALIDO']:
                        estado_val = 'BORRADOR'

                    def get_float(val, default=0):
                        try:
                            if pd.isna(val) or val == '':
                                return default
                            return float(val)
                        except:
                            return default

                    factor1_val = get_float(row['factor1'])
                    factor2_val = get_float(row['factor2'])
                    factor3_val = get_float(row['factor3'])
                    factor4_val = get_float(row['factor4'])
                    factor5_val = get_float(row['factor5'])
                    factor6_val = get_float(row['factor6'])
                    factor7_val = get_float(row['factor7'])
                    factor8_val = get_float(row['factor8'])
                    factor9_val = get_float(row['factor9'])
                    factor10_val = get_float(row['factor10'])
                    factor11_val = get_float(row['factor11'])
                    factor12_val = get_float(row['factor12'])
                    factor13_val = get_float(row['factor13'])
                    factor14_val = get_float(row['factor14'])
                    factor15_val = get_float(row['factor15'])
                    factor16_val = get_float(row['factor16'])
                    factor17_val = get_float(row['factor17'])
                    factor18_val = get_float(row['factor18'])
                    factor19_val = get_float(row['factor19'])
                    factor20_val = get_float(row['factor20'])
                    factor21_val = get_float(row['factor21'])
                    factor22_val = get_float(row['factor22'])
                    factor23_val = get_float(row['factor23'])
                    factor24_val = get_float(row['factor24'])
                    factor25_val = get_float(row['factor25'])
                    factor26_val = get_float(row['factor26'])

                    # Chequea la suma de los factores
                    total_factores = factor1_val + factor2_val + factor3_val + factor4_val + factor5_val + factor6_val + factor7_val + factor8_val + factor9_val + factor10_val + factor11_val + factor12_val + factor13_val + factor14_val + factor15_val + factor16_val + factor17_val + factor18_val + factor19_val + factor20_val + factor21_val + factor22_val + factor23_val + factor24_val + factor25_val + factor26_val
                    if total_factores > 1:
                        estado_val = 'INVALIDO'
                    else:
                        # Normaliza factores cuando la suma es = 1 if total > 0
                        if total_factores > 0:
                            factor1_val /= total_factores
                            factor2_val /= total_factores
                            factor3_val /= total_factores
                            factor4_val /= total_factores
                            factor5_val /= total_factores
                            factor6_val /= total_factores
                            factor7_val /= total_factores
                            factor8_val /= total_factores
                            factor9_val /= total_factores
                            factor10_val /= total_factores
                            factor11_val /= total_factores
                            factor12_val /= total_factores
                            factor13_val /= total_factores
                            factor14_val /= total_factores
                            factor15_val /= total_factores
                            factor16_val /= total_factores
                            factor17_val /= total_factores
                            factor18_val /= total_factores
                            factor19_val /= total_factores
                            factor20_val /= total_factores
                            factor21_val /= total_factores
                            factor22_val /= total_factores
                            factor23_val /= total_factores
                            factor24_val /= total_factores
                            factor25_val /= total_factores
                            factor26_val /= total_factores

                    calif = Calificacion(
                        tipo_mercado=tipo_mercado_val,
                        origen=origen_val,
                        mercado=mercado_val,
                        instrumento=instrumento_val,
                        secuencia=secuencia_val,
                        numero_dividendo=numero_dividendo_val,
                        tipo_sociedad=tipo_sociedad_val,
                        fecha=fecha_val,
                        valor_historico=valor_historico_val,
                        ejercicio=ejercicio_val,
                        descripcion=descripcion_val,
                        isfut_casilla=isfut_casilla_val,
                        factor_actualizacion=factor_actualizacion_val,
                        factor1=factor1_val,
                        factor2=factor2_val,
                        factor3=factor3_val,
                        factor4=factor4_val,
                        factor5=factor5_val,
                        factor6=factor6_val,
                        factor7=factor7_val,
                        factor8=factor8_val,
                        factor9=factor9_val,
                        factor10=factor10_val,
                        factor11=factor11_val,
                        factor12=factor12_val,
                        factor13=factor13_val,
                        factor14=factor14_val,
                        factor15=factor15_val,
                        factor16=factor16_val,
                        factor17=factor17_val,
                        factor18=factor18_val,
                        factor19=factor19_val,
                        factor20=factor20_val,
                        factor21=factor21_val,
                        factor22=factor22_val,
                        factor23=factor23_val,
                        factor24=factor24_val,
                        factor25=factor25_val,
                        factor26=factor26_val,
                        estado=estado_val
                    )
                    calificaciones.append(calif)
                except Exception as e:
                    errors.append(f"Error en fila {_+1} del archivo {file.name}: {str(e)}")
                    continue

            Calificacion.objects.bulk_create(calificaciones)
            total_calificaciones += len(calificaciones)
        except Exception as e:
            errors.append(f"Error al procesar el archivo {file.name}: {str(e)}")
            continue

    if errors:
        return JsonResponse({'success': False, 'error': '; '.join(errors)})
    else:
        return JsonResponse({'success': True, 'count': total_calificaciones})

@method_decorator(login_required, name='dispatch')
class BulkUploadView(FormView):
    template_name = "calificaciones/bulk_upload.html"
    form_class = BulkUploadForm
    success_url = reverse_lazy("calif_list")

    def form_valid(self, form):
        files = self.request.FILES.getlist('file')
        total_calificaciones = 0
        for file in files:
            try:
                if file.name.endswith('.csv'):
                    df = pd.read_csv(file)
                elif file.name.endswith(('.xlsx', '.xls')):
                    df = pd.read_excel(file)
                elif file.name.endswith('.pdf'):
                    with pdfplumber.open(file) as pdf:
                        page = pdf.pages[0]
                        table = page.extract_table()
                        if table:
                            df = pd.DataFrame(table[1:], columns=table[0])
                        else:
                            messages.error(self.request, f"No se encontró tabla en el PDF {file.name}.")
                            continue
                else:
                    messages.error(self.request, f"Formato de archivo no soportado para {file.name}. Usa CSV, XLSX o PDF.")
                    continue

                # Solo acepta columnas con nombre
                required_columns = ['tipo_mercado', 'origen', 'mercado', 'instrumento', 'secuencia', 'numero_dividendo', 'tipo_sociedad', 'fecha', 'valor_historico', 'ejercicio', 'descripcion', 'isfut_casilla', 'factor_actualizacion', 'factor1', 'factor2', 'factor3', 'factor4', 'factor5', 'factor6', 'factor7', 'factor8', 'factor9', 'factor10', 'factor11', 'factor12', 'factor13', 'factor14', 'factor15', 'factor16', 'factor17', 'factor18', 'factor19', 'factor20', 'factor21', 'factor22', 'factor23', 'factor24', 'factor25', 'factor26', 'estado']
                if not all(col in df.columns for col in required_columns):
                    messages.error(self.request, f"El archivo {file.name} no tiene las columnas requeridas. Descarga la plantilla y utilicela como base.")
                    continue

                calificaciones = []
                for _, row in df.iterrows():
                    try:
                        tipo_mercado_val = str(row['tipo_mercado']).strip() if row['tipo_mercado'] else 'ACCIONES'
                        origen_val = str(row['origen']).strip() if row['origen'] else 'CORREDORA'
                        mercado_val = str(row['mercado']).strip()
                        instrumento_val = str(row['instrumento']).strip()
                        secuencia_val = int(float(row['secuencia'])) if row['secuencia'] and str(row['secuencia']).strip() else 0
                        numero_dividendo_val = int(float(row['numero_dividendo'])) if row['numero_dividendo'] and str(row['numero_dividendo']).strip() else 0
                        tipo_sociedad_val = str(row['tipo_sociedad']).strip() if row['tipo_sociedad'] else 'A'
                        if tipo_sociedad_val not in ['A', 'C']:
                            tipo_sociedad_val = 'A'
                        fecha_val = pd.to_datetime(row['fecha'], errors='coerce') if row['fecha'] else datetime(2023, 1, 1)
                        if pd.isna(fecha_val):
                            fecha_val = datetime(2023, 1, 1)
                        fecha_val = fecha_val.date()
                        valor_historico_val = float(row['valor_historico']) if row['valor_historico'] else 0.0
                        ejercicio_val = pd.to_datetime(row['ejercicio'], errors='coerce') if row['ejercicio'] else datetime(2023, 1, 1)
                        if pd.isna(ejercicio_val):
                            ejercicio_val = datetime(2023, 1, 1)
                        ejercicio_val = ejercicio_val.date()
                        descripcion_val = str(row['descripcion']).strip() if row['descripcion'] else ''
                        # Convertir isfut_casilla a booleano
                        isfut_casilla_val = str(row['isfut_casilla']).strip().lower() if row['isfut_casilla'] else ''
                        if isfut_casilla_val in ['true', 'verdadero', '1', 'yes', 'si']:
                            isfut_casilla_val = True
                        elif isfut_casilla_val in ['false', 'falso', '0', 'no']:
                            isfut_casilla_val = False
                        else:
                            isfut_casilla_val = False  # Valor predeterminado: Falso si no se reconoce
                        factor_actualizacion_val = float(row['factor_actualizacion']) if row['factor_actualizacion'] else 1.0
                        estado_val = str(row.get('estado', 'BORRADOR')).strip() or 'BORRADOR'
                        if estado_val not in ['BORRADOR', 'EN PROCESO', 'VIGENTE', 'ANULADA', 'INVALIDO']:
                            estado_val = 'BORRADOR'

                        def get_float(val, default=0):
                            try:
                                if pd.isna(val) or val == '':
                                    return default
                                return float(val)
                            except:
                                return default

                        factor1_val = get_float(row['factor1'])
                        factor2_val = get_float(row['factor2'])
                        factor3_val = get_float(row['factor3'])
                        factor4_val = get_float(row['factor4'])
                        factor5_val = get_float(row['factor5'])
                        factor6_val = get_float(row['factor6'])
                        factor7_val = get_float(row['factor7'])
                        factor8_val = get_float(row['factor8'])
                        factor9_val = get_float(row['factor9'])
                        factor10_val = get_float(row['factor10'])
                        factor11_val = get_float(row['factor11'])
                        factor12_val = get_float(row['factor12'])
                        factor13_val = get_float(row['factor13'])
                        factor14_val = get_float(row['factor14'])
                        factor15_val = get_float(row['factor15'])
                        factor16_val = get_float(row['factor16'])
                        factor17_val = get_float(row['factor17'])
                        factor18_val = get_float(row['factor18'])
                        factor19_val = get_float(row['factor19'])
                        factor20_val = get_float(row['factor20'])
                        factor21_val = get_float(row['factor21'])
                        factor22_val = get_float(row['factor22'])
                        factor23_val = get_float(row['factor23'])
                        factor24_val = get_float(row['factor24'])
                        factor25_val = get_float(row['factor25'])
                        factor26_val = get_float(row['factor26'])

                        # Chequea la suma de los factores
                        total_factores = factor1_val + factor2_val + factor3_val + factor4_val + factor5_val + factor6_val + factor7_val + factor8_val + factor9_val + factor10_val + factor11_val + factor12_val + factor13_val + factor14_val + factor15_val + factor16_val + factor17_val + factor18_val + factor19_val + factor20_val + factor21_val + factor22_val + factor23_val + factor24_val + factor25_val + factor26_val
                        if total_factores > 1:
                            estado_val = 'INVALIDO'
                        else:
                            # Normaliza factores cuando la suma es = 1 if total > 0
                            if total_factores > 0:
                                factor1_val /= total_factores
                                factor2_val /= total_factores
                                factor3_val /= total_factores
                                factor4_val /= total_factores
                                factor5_val /= total_factores
                                factor6_val /= total_factores
                                factor7_val /= total_factores
                                factor8_val /= total_factores
                                factor9_val /= total_factores
                                factor10_val /= total_factores
                                factor11_val /= total_factores
                                factor12_val /= total_factores
                                factor13_val /= total_factores
                                factor14_val /= total_factores
                                factor15_val /= total_factores
                                factor16_val /= total_factores
                                factor17_val /= total_factores
                                factor18_val /= total_factores
                                factor19_val /= total_factores
                                factor20_val /= total_factores
                                factor21_val /= total_factores
                                factor22_val /= total_factores
                                factor23_val /= total_factores
                                factor24_val /= total_factores
                                factor25_val /= total_factores
                                factor26_val /= total_factores

                        calif = Calificacion(
                            tipo_mercado=tipo_mercado_val,
                            origen=origen_val,
                            mercado=mercado_val,
                            instrumento=instrumento_val,
                            secuencia=secuencia_val,
                            numero_dividendo=numero_dividendo_val,
                            tipo_sociedad=tipo_sociedad_val,
                            fecha=fecha_val,
                            valor_historico=valor_historico_val,
                            ejercicio=ejercicio_val,
                            descripcion=descripcion_val,
                            isfut_casilla=isfut_casilla_val,
                            factor_actualizacion=factor_actualizacion_val,
                            factor1=factor1_val,
                            factor2=factor2_val,
                            factor3=factor3_val,
                            factor4=factor4_val,
                            factor5=factor5_val,
                            factor6=factor6_val,
                            factor7=factor7_val,
                            factor8=factor8_val,
                            factor9=factor9_val,
                            factor10=factor10_val,
                            factor11=factor11_val,
                            factor12=factor12_val,
                            factor13=factor13_val,
                            factor14=factor14_val,
                            factor15=factor15_val,
                            factor16=factor16_val,
                            factor17=factor17_val,
                            factor18=factor18_val,
                            factor19=factor19_val,
                            factor20=factor20_val,
                            factor21=factor21_val,
                            factor22=factor22_val,
                            factor23=factor23_val,
                            factor24=factor24_val,
                            factor25=factor25_val,
                            factor26=factor26_val,
                            estado=estado_val
                        )
                        calificaciones.append(calif)
                    except Exception as e:
                        messages.warning(self.request, f"Error en fila {_+1} del archivo {file.name}: {str(e)}")
                        continue

                Calificacion.objects.bulk_create(calificaciones)
                total_calificaciones += len(calificaciones)
                messages.success(self.request, f"Se cargaron {len(calificaciones)} calificaciones del archivo {file.name}.")
            except Exception as e:
                messages.error(self.request, f"Error al procesar el archivo {file.name}: {str(e)}")
                continue

        if total_calificaciones > 0:
            messages.success(self.request, f"Total calificaciones cargadas: {total_calificaciones}.")
        else:
            messages.warning(self.request, "No se cargaron calificaciones.")

        return super().form_valid(form)
